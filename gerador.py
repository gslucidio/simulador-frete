"""
Dashboard — Modelagem de Frete (v5)
Suba este arquivo + Modelagem_Frete_v5.xlsx + requirements.txt no GitHub
e faça deploy em share.streamlit.io
"""

import io
import datetime
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from openpyxl import load_workbook
from openpyxl.cell import MergedCell

st.set_page_config(page_title="Modelo de Frete", page_icon="🚛", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');
html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
h1, h2, h3 { font-family: 'IBM Plex Mono', monospace; }
.kpi { background:#0f1117; border:1px solid #2a2d3a; border-left:3px solid #00d4aa;
       border-radius:4px; padding:14px 18px; margin-bottom:8px; }
.kpi-neg  { border-left-color:#e05c5c !important; }
.kpi-warn { border-left-color:#f5a623 !important; }
.kpi-label { color:#8b8fa8; font-size:11px; text-transform:uppercase; letter-spacing:1px; }
.kpi-value { color:#fff; font-size:22px; font-family:'IBM Plex Mono',monospace; font-weight:600; }
.kpi-sub   { color:#8b8fa8; font-size:12px; margin-top:2px; }
.sec { font-family:'IBM Plex Mono',monospace; font-size:11px; font-weight:600; color:#00d4aa;
       letter-spacing:2px; text-transform:uppercase; border-bottom:1px solid #2a2d3a;
       padding-bottom:6px; margin:18px 0 10px 0; }
.warn-box { background:#1a1200; border:1px solid #f5a623; border-radius:4px;
            padding:10px 14px; margin-bottom:10px; color:#f5a623; font-size:13px; }
.stSidebar { background-color:#0a0c12; }
[data-testid="stSidebar"] label { color:#c0c4d6 !important; font-size:13px; }
div[data-testid="stDownloadButton"] button {
    background:#00d4aa; color:#0a0c12; font-family:'IBM Plex Mono',monospace;
    font-weight:600; border:none; border-radius:4px; padding:10px 24px; width:100%; }
</style>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
# GERADOR EXCEL
# ════════════════════════════════════════════════════════════════════════════

def gerar_excel(template_bytes, p):
    N = p["n_dias"]; ORIG = N - 21; WIND = ORIG + 5; LR = N + 4
    START = datetime.datetime(2026, 6, 1)

    wb   = load_workbook(io.BytesIO(template_bytes))
    ws_f = wb["Fundo"]; ws_c = wb["Custos"]; ws_d = wb["Dashboard"]

    def sc(ws, row, col, val):
        cell = ws.cell(row, col)
        if not isinstance(cell, MergedCell): cell.value = val

    def clear_rows(ws, from_row):
        for r in range(from_row, ws.max_row + 1):
            for col in range(1, ws.max_column + 1): sc(ws, r, col, None)

    # Hipóteses do Dashboard
    ws_d["I3"].value = p["pc"];      ws_d["I4"].value = p["pv"]
    ws_d["I8"].value = p["orig"];    ws_d["I9"].value = p["p18"]
    ws_d["I10"].value = p["p20"];    ws_d["I11"].value = p["p22"]
    ws_d["D14"].value = p["pl"]
    ws_d["C10"].value = p["pct_sr"]; ws_d["C11"].value = p["pct_mz"]
    ws_d["C12"].value = round(1 - p["pct_sr"] - p["pct_mz"], 4)
    ws_d["D17"].value = p["cdi"];    ws_d["D18"].value = p["sr_sp"]
    ws_d["D19"].value = p["mz_sp"]
    ws_d["I14"].value = p["cons"];   ws_d["I15"].value = p["gest"]
    ws_d["J15"].value = p["piso_g"]; ws_d["I16"].value = p["adm"]
    ws_d["J16"].value = p["piso_a"]

    # Dashboard: cells dinâmicas
    ws_d["E5"].value  = f"=Fundo!AD5+Fundo!AD{LR}"
    ws_d["E6"].value  = f"=Fundo!AJ5+Fundo!AJ{LR}"
    ws_d["E7"].value  = f"=Fundo!AN5+Fundo!AN{LR}"
    ws_d["C23"].value = f"=Fundo!U{LR}/Dashboard!D14"

    # IRRs
    ws_f["X2"].value  = f"=IRR(X5:X{LR},0.001)"
    ws_f["AD2"].value = f"=IRR(AD5:AD{LR},0.001)"
    ws_f["AJ2"].value = f"=IRR(AJ5:AJ{LR},0.001)"
    ws_f["AN2"].value = f"=IRR(AN5:AN{LR},0.001)"

    clear_rows(ws_f, 5); clear_rows(ws_c, 5)

    for n in range(1, N + 1):
        r = n + 4; pr = r - 1
        first = (n == 1); last = (n == N)
        wind = (r == WIND); q_eim = (r >= WIND)

        sc(ws_f,r,2, "=Dashboard!D14" if first else f"=B{pr}*(((1+Dashboard!$C$20)^(1/30)))")
        sc(ws_f,r,3, n)
        sc(ws_f,r,4, START if first else f"=D{pr}+1")

        if first:
            sc(ws_f,r,5,"=Q5*Dashboard!$I$9"); sc(ws_f,r,9,"=Q5*Dashboard!$I$10"); sc(ws_f,r,13,"=Q5*Dashboard!$I$11")
        elif wind:
            sc(ws_f,r,5,0); sc(ws_f,r,9,0); sc(ws_f,r,13,0)
        else:
            sc(ws_f,r,5,f"=E{pr}"); sc(ws_f,r,9,f"=I{pr}"); sc(ws_f,r,13,f"=M{pr}")

        sc(ws_f,r,6, f"=E{r}*Dashboard!$I$3"); sc(ws_f,r,10,f"=I{r}*Dashboard!$I$3"); sc(ws_f,r,14,f"=M{r}*Dashboard!$I$3")
        sc(ws_f,r,7,  0 if r<22 else f"=E{r-17}*Dashboard!$I$4")
        sc(ws_f,r,11, 0 if r<24 else f"=I{r-19}*Dashboard!$I$4")
        sc(ws_f,r,15, 0 if r<26 else f"=M{r-21}*Dashboard!$I$4")
        sc(ws_f,r,17, "=Dashboard!I8" if first else (f"=E{r}+I{r}+M{r}" if q_eim else f"=Q{pr}"))
        sc(ws_f,r,18, f"=F{r}+J{r}+N{r}"); sc(ws_f,r,19, f"=G{r}+K{r}+O{r}")
        sc(ws_f,r,20, 0 if r<22 else f"=S{r}-R{r-17}")
        sc(ws_f,r,22, 0 if first else f"=U{pr}*(((1+Dashboard!$C$17)^(1/30))-1)")

        if first:  sc(ws_f,r,21, f"=U4-R{r}+S{r}+V{r}-Custos!D{r}")
        elif n==2: sc(ws_f,r,21, f"=U{pr}-R{r}+S{r}+V{r}-Custos!D{r}-AB{pr}-AH{pr}-AM{pr}")
        else:      sc(ws_f,r,21, f"=U{pr}-R{r}+S{r}+V{r}-Custos!D{r}+Z{pr}+AF{pr}+AL{pr}-AB{pr}-AH{pr}-AM{pr}")

        sc(ws_f,r,23, f"=U{r}+R{r}"); sc(ws_f,r,24, f"=-R{r}+S{r}-Custos!D{r}")
        sc(ws_f,r,26, "=Z4" if first else 0)
        sc(ws_f,r,27, f"=AC{pr}*(((1+Dashboard!$C$20)^(1/30))-1)")
        sc(ws_f,r,28, f"=AC{r}" if last else 0)
        sc(ws_f,r,29, f"=Z{r}" if first else f"=AC{pr}+AA{r}-AB{pr}")
        sc(ws_f,r,30, f"=-Z{r}+AB{r}")
        sc(ws_f,r,32, "=AF4" if first else 0)
        sc(ws_f,r,33, f"=AI{pr}*(((1+Dashboard!$C$21)^(1/30))-1)")
        sc(ws_f,r,34, f"=AI{r}" if last else 0)
        sc(ws_f,r,35, f"=AF{r}" if first else f"=AI{pr}+AG{r}-AH{pr}")
        sc(ws_f,r,36, f"=-AF{r}+AH{r}")
        sc(ws_f,r,38, "=AL4" if first else 0)
        sc(ws_f,r,39, f"=U{r}-AB{r}-AH{r}" if last else 0)
        sc(ws_f,r,40, f"=-AL{r}+AM{r}")

        sc(ws_c,r,2, n)
        sc(ws_c,r,3, START if first else f"=C{pr}+1")
        sc(ws_c,r,5, "=Dashboard!$J$16/30" if first else f"=MAX(Dashboard!$J$16,(Dashboard!$I$16*Fundo!W{pr})/360)/30")
        sc(ws_c,r,6, "=Dashboard!$J$15/30" if first else f"=MAX(Dashboard!$J$15,(Dashboard!$I$15*Fundo!W{pr})/360)/30")
        sc(ws_c,r,7, 0 if first else f"=Fundo!R{pr}*Dashboard!$I$14")
        sc(ws_c,r,8, 0)   # Performance zerada
        sc(ws_c,r,9, 10000 if first else None)
        sc(ws_c,r,4, f"=E{r}+F{r}+G{r}+H{r}+I{r}")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# SIMULAÇÃO NUMÉRICA (para os gráficos)
# ════════════════════════════════════════════════════════════════════════════

def simular(p):
    N = p["n_dias"]; ORIG = N - 21
    pc = p["pc"]; pv = p["pv"]
    p18 = p["p18"]; p20 = p["p20"]; p22 = p["p22"]
    pl = p["pl"]
    sr0 = pl * p["pct_sr"]
    mz0 = pl * p["pct_mz"]
    jr0 = pl * round(1 - p["pct_sr"] - p["pct_mz"], 4)

    cdi_m = (1 + p["cdi"]) ** (1/12) - 1
    sr_m  = (1 + p["sr_sp"]) ** (1/12) - 1
    mz_m  = (1 + p["mz_sp"]) ** (1/12) - 1
    C20   = (1 + cdi_m) * (1 + sr_m) - 1
    C21   = (1 + cdi_m) * (1 + mz_m) - 1
    cdi_d = (1 + cdi_m) ** (1/30) - 1
    sr_d  = (1 + C20)   ** (1/30) - 1
    mz_d  = (1 + C21)   ** (1/30) - 1

    piso_a = p["piso_a"] / 30
    piso_g = p["piso_g"] / 30

    U = pl; sr = sr0; mz = mz0
    E_h, I_h, M_h, R_h = [], [], [], []
    dias, U_h, W_h, sr_h, mz_h, jr_h = [], [], [], [], [], []
    c_adm = c_gest = c_cons = 0

    for n in range(1, N + 1):
        first = (n == 1)
        Q  = p["orig"] if n <= ORIG else 0
        E_ = Q*p18; I_ = Q*p20; M_ = Q*p22
        E_h.append(E_); I_h.append(I_); M_h.append(M_)
        desemb = Q * pc; R_h.append(desemb)

        G  = E_h[n-18]*pv if n >= 18 else 0
        K  = I_h[n-20]*pv if n >= 20 else 0
        O_ = M_h[n-22]*pv if n >= 22 else 0
        receb = G + K + O_

        V = U * cdi_d if not first else 0
        W = U + desemb

        adm  = max(piso_a, (p["adm"]  * W / 360) / 30) if not first else piso_a
        gest = max(piso_g, (p["gest"] * W / 360) / 30) if not first else piso_g
        cons = R_h[n-2] * p["cons"] if not first else 0
        outros = 10000 if first else 0
        custo = adm + gest + cons + outros   # sem performance

        c_adm += adm; c_gest += gest; c_cons += cons

        sr += sr * sr_d
        mz += mz * mz_d
        U   = U - desemb + receb + V - custo

        dias.append(n); U_h.append(U); W_h.append(W)
        sr_h.append(sr); mz_h.append(mz); jr_h.append(U - sr - mz)

    return {
        "df": pd.DataFrame({"dia": dias, "caixa": U_h, "pl_fundo": W_h,
                            "pl_sr": sr_h, "pl_mz": mz_h, "pl_jr": jr_h}),
        "U_final": U, "am_sr": sr, "am_mz": mz, "am_jr": U - sr - mz,
        "sr0": sr0, "mz0": mz0, "jr0": jr0,
        "c_adm": c_adm, "c_gest": c_gest, "c_cons": c_cons,
        "margem_bruta": ORIG * p["orig"] * (pv - pc),
        "carteira_peak": p["orig"] * pc * (18*p18 + 20*p20 + 22*p22),
    }


# ════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## 🚛 Modelo de Frete")
    st.markdown("---")

    st.markdown('<div class="sec">Período</div>', unsafe_allow_html=True)
    n_dias = st.number_input("Número de dias", 50, 1800, 395, 10)

    st.markdown('<div class="sec">Operação</div>', unsafe_allow_html=True)
    pc   = st.number_input("Preço de compra (R$)", value=1678, step=10)
    pv   = st.number_input("Preço de venda (R$)",  value=1750, step=10)
    orig = st.number_input("Originações por dia",  value=30,   step=1)

    st.markdown('<div class="sec">Prazos de recebimento</div>', unsafe_allow_html=True)
    p18 = st.slider("% 18 dias", 0.0, 1.0, 0.30, 0.05, format="%.0%%")
    p20 = st.slider("% 20 dias", 0.0, 1.0, 0.30, 0.05, format="%.0%%")
    p22 = st.slider("% 22 dias", 0.0, 1.0, 0.40, 0.05, format="%.0%%")
    if abs(p18 + p20 + p22 - 1.0) > 0.01:
        st.warning(f"⚠️ Soma = {p18+p20+p22:.0%} (deve ser 100%)")

    st.markdown('<div class="sec">Estrutura do fundo</div>', unsafe_allow_html=True)
    pl     = st.number_input("PL Total (R$)", value=1_000_000, step=50_000)
    pct_sr = st.slider("% Sênior",   0.0, 1.0, 0.60, 0.05, format="%.0%%")
    pct_mz = st.slider("% Mezanino", 0.0, 1.0, 0.10, 0.05, format="%.0%%")
    pct_jr = round(1 - pct_sr - pct_mz, 2)
    st.caption(f"Júnior (residual): {pct_jr:.0%}")

    st.markdown('<div class="sec">Taxas</div>', unsafe_allow_html=True)
    cdi   = st.number_input("CDI a.a. (%)",      value=14.5, step=0.5) / 100
    sr_sp = st.number_input("Sênior CDI+ (%)",   value=3.0,  step=0.5) / 100
    mz_sp = st.number_input("Mezanino CDI+ (%)", value=5.0,  step=0.5) / 100

    st.markdown('<div class="sec">Custos</div>', unsafe_allow_html=True)
    cons   = st.number_input("Consultoria (% desembolso)", value=1.0,   step=0.1) / 100
    gest   = st.number_input("Gestão a.a. (%)",            value=0.5,   step=0.1) / 100
    piso_g = st.number_input("Piso Gestão R$/mês",         value=15000, step=1000)
    adm    = st.number_input("Adm a.a. (%)",               value=0.2,   step=0.1) / 100
    piso_a = st.number_input("Piso Adm R$/mês",            value=15000, step=1000)

    st.markdown("---")
    tpl_file = st.file_uploader("Template .xlsx", type=["xlsx"])


# ════════════════════════════════════════════════════════════════════════════
# PARÂMETROS E SIMULAÇÃO
# ════════════════════════════════════════════════════════════════════════════

params = dict(
    n_dias=n_dias, pc=pc, pv=pv, orig=orig,
    p18=p18, p20=p20, p22=p22,
    pl=pl, pct_sr=pct_sr, pct_mz=pct_mz,
    cdi=cdi, sr_sp=sr_sp, mz_sp=mz_sp,
    cons=cons, gest=gest, piso_g=piso_g,
    adm=adm, piso_a=piso_a,
)

res  = simular(params)
df   = res["df"]

carteira_peak = res["carteira_peak"]
util          = carteira_peak / pl
moic_f        = res["U_final"] / pl
moic_sr       = res["am_sr"] / res["sr0"]
moic_mz       = res["am_mz"] / res["mz0"]
moic_jr       = res["am_jr"] / res["jr0"] if res["jr0"] > 0 else 0
custo_total   = res["c_adm"] + res["c_gest"] + res["c_cons"] + 10000
margem        = res["margem_bruta"]
ops_max       = int(pl / (pc * (18*p18 + 20*p20 + 22*p22))) if (18*p18+20*p20+22*p22) > 0 else 0


# ════════════════════════════════════════════════════════════════════════════
# LAYOUT
# ════════════════════════════════════════════════════════════════════════════

st.markdown("# Modelagem de Frete")
st.markdown(f"**{n_dias} dias** · Originação até dia {n_dias-21} · {orig} ops/dia · PL R$ {pl/1e6:.2f}M")
st.markdown("---")

# Alertas
if util > 1.0:
    st.markdown(
        f'<div class="warn-box">⚠️ <b>Fundo superdimensionado:</b> carteira no pico = R$ {carteira_peak:,.0f} ({util:.1%} do PL). '
        f'Reduza para {ops_max} ops/dia para evitar caixa negativo.</div>',
        unsafe_allow_html=True)
if moic_jr < 1.0:
    st.markdown(
        f'<div class="warn-box">⚠️ <b>Júnior com retorno negativo:</b> custos (R$ {custo_total:,.0f}) vs margem bruta (R$ {margem:,.0f}). '
        f'Revise volume de originação, prazos ou estrutura de custos.</div>',
        unsafe_allow_html=True)

# KPIs
def kpi(col, label, val, sub="", neg=False, warn=False):
    cls = "kpi-neg" if neg else ("kpi-warn" if warn else "")
    col.markdown(
        f'<div class="kpi {cls}">'
        f'<div class="kpi-label">{label}</div>'
        f'<div class="kpi-value">{val}</div>'
        f'<div class="kpi-sub">{sub}</div>'
        f'</div>', unsafe_allow_html=True)

c1, c2, c3, c4, c5, c6 = st.columns(6)
kpi(c1, "Margem/op",   f"R$ {pv-pc:,.0f}",     f"Antecipação: {pv/pc-1:.2%}")
kpi(c2, "Utilização",  f"{util:.1%}",           f"Pico: R$ {carteira_peak/1e3:.0f}k", warn=util>1.0)
kpi(c3, "MOIC Fundo",  f"{moic_f:.3f}x",        f"Final: R$ {res['U_final']/1e3:.0f}k", neg=moic_f<1)
kpi(c4, "MOIC Sênior", f"{moic_sr:.3f}x",       f"CDI+{sr_sp:.0%}", neg=moic_sr<1)
kpi(c5, "MOIC Jr",     f"{moic_jr:.3f}x",       f"Entrou R$ {res['jr0']/1e3:.0f}k", neg=moic_jr<1)
kpi(c6, "Ops Total",   f"{(n_dias-21)*orig:,}",  f"R$ {(n_dias-21)*orig*(pv-pc)/1e3:.0f}k margem")
st.markdown("")

# Estilo dos gráficos
PLOT = dict(
    plot_bgcolor="#0a0c12", paper_bgcolor="#0a0c12",
    font=dict(color="#c0c4d6", family="IBM Plex Mono"),
    margin=dict(l=0, r=0, t=10, b=0),
    legend=dict(bgcolor="#0f1117", bordercolor="#2a2d3a", borderwidth=1))

# Gráfico 1: PL por cota
st.markdown('<div class="sec">Evolução do PL por cota</div>', unsafe_allow_html=True)
fig1 = go.Figure()
fig1.add_trace(go.Scatter(x=df.dia, y=df.pl_sr/1e3, name="Sênior",
    line=dict(color="#4a9eff", width=2)))
fig1.add_trace(go.Scatter(x=df.dia, y=df.pl_mz/1e3, name="Mezanino",
    line=dict(color="#f5a623", width=2)))
fig1.add_trace(go.Scatter(x=df.dia, y=df.pl_jr/1e3, name="Júnior",
    line=dict(color="#e05c5c", width=2),
    fill="tozeroy", fillcolor="rgba(224,92,92,0.08)"))
fig1.add_hline(y=0, line_dash="dash", line_color="#444", line_width=1)
fig1.update_layout(**PLOT, height=300,
    xaxis=dict(title="Dia", gridcolor="#1a1d2a"),
    yaxis=dict(title="R$ mil", gridcolor="#1a1d2a"))
st.plotly_chart(fig1, use_container_width=True)

# Gráficos 2 e 3 lado a lado
ca, cb = st.columns(2)

with ca:
    st.markdown('<div class="sec">Caixa (U)</div>', unsafe_allow_html=True)
    fig2 = go.Figure()
    fig2.add_trace(go.Scatter(x=df.dia, y=df.caixa/1e3, name="Caixa",
        fill="tozeroy", fillcolor="rgba(0,212,170,0.08)",
        line=dict(color="#00d4aa", width=2)))
    fig2.add_hline(y=0, line_dash="dash", line_color="#e05c5c", line_width=1)
    fig2.update_layout(**PLOT, height=270, showlegend=False,
        xaxis=dict(title="Dia", gridcolor="#1a1d2a"),
        yaxis=dict(title="R$ mil", gridcolor="#1a1d2a"))
    st.plotly_chart(fig2, use_container_width=True)

with cb:
    st.markdown('<div class="sec">Decomposição de custos</div>', unsafe_allow_html=True)
    labels = ["Adm", "Gestão", "Consultoria", "Outros"]
    vals   = [res["c_adm"], res["c_gest"], res["c_cons"], 10000]
    colors = ["#4a9eff", "#00d4aa", "#f5a623", "#888"]
    fig3 = go.Figure(go.Bar(
        x=labels, y=vals, marker_color=colors,
        text=[f"R${v/1e3:.0f}k" for v in vals],
        textposition="outside",
        textfont=dict(family="IBM Plex Mono", size=11)))
    fig3.add_hline(y=margem, line_dash="dash", line_color="#00d4aa",
        annotation_text=f"Margem bruta ops: R$ {margem/1e3:.0f}k",
        annotation_font_color="#00d4aa",
        annotation_position="top right")
    fig3.update_layout(**PLOT, height=270, showlegend=False,
        xaxis=dict(gridcolor="#1a1d2a"),
        yaxis=dict(title="R$", gridcolor="#1a1d2a"))
    st.plotly_chart(fig3, use_container_width=True)

# Tabela de liquidação
st.markdown('<div class="sec">Resumo de liquidação</div>', unsafe_allow_html=True)
tbl = pd.DataFrame({
    "Cota":      ["Sênior", "Mezanino", "Júnior", "Total Fundo"],
    "Entrada":   [f"R$ {res['sr0']:,.0f}", f"R$ {res['mz0']:,.0f}",
                  f"R$ {res['jr0']:,.0f}", f"R$ {pl:,.0f}"],
    "Saída":     [f"R$ {res['am_sr']:,.0f}", f"R$ {res['am_mz']:,.0f}",
                  f"R$ {res['am_jr']:,.0f}", f"R$ {res['U_final']:,.0f}"],
    "MOIC":      [f"{moic_sr:.3f}x", f"{moic_mz:.3f}x",
                  f"{moic_jr:.3f}x", f"{moic_f:.3f}x"],
    "Ganho":     [f"R$ {res['am_sr']-res['sr0']:+,.0f}",
                  f"R$ {res['am_mz']-res['mz0']:+,.0f}",
                  f"R$ {res['am_jr']-res['jr0']:+,.0f}",
                  f"R$ {res['U_final']-pl:+,.0f}"],
})
st.dataframe(tbl, use_container_width=True, hide_index=True)

# Download Excel
st.markdown("---")
st.markdown('<div class="sec">Exportar modelo</div>', unsafe_allow_html=True)

tpl_bytes = None
if tpl_file:
    tpl_bytes = tpl_file.read()
else:
    try:
        with open("Modelagem_Frete_v5.xlsx", "rb") as f:
            tpl_bytes = f.read()
    except FileNotFoundError:
        st.info("📁 Suba o `Modelagem_Frete_v5.xlsx` junto com o app no GitHub, ou use o upload acima na sidebar.")

if tpl_bytes:
    with st.spinner("Gerando Excel..."):
        xl = gerar_excel(tpl_bytes, params)
    st.download_button(
        label=f"⬇ Baixar Modelagem_Frete_{n_dias}d.xlsx",
        data=xl,
        file_name=f"Modelagem_Frete_{n_dias}d.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with st.expander("Ver dados diários da simulação"):
    df_s = df.copy()
    for col in ["caixa", "pl_fundo", "pl_sr", "pl_mz", "pl_jr"]:
        df_s[col] = df_s[col].apply(lambda x: f"R$ {x:,.0f}")
    df_s.columns = ["Dia", "Caixa", "PL Fundo", "Sênior", "Mezanino", "Júnior"]
    st.dataframe(df_s, use_container_width=True, height=280)
